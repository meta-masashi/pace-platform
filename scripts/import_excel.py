#!/usr/bin/env python3
"""
PACE Platform - Excel Import Script  (バックログ #2)
======================================================
実際の Excel ファイルを読み込み Supabase へ upsert する。

使い方:
    python import_excel.py --dry-run            # プレビュー（書き込みなし）
    python import_excel.py                      # 全ファイルをインポート
    python import_excel.py --only assessment    # アセスメントのみ
    python import_excel.py --only exercises     # エクササイズのみ
    python import_excel.py --only rtp           # RTP ノードのみ
    python import_excel.py --only mc            # MC トラッキングのみ

防壁1 準拠: モック実装禁止 — 実際の Excel を読んで動作する。
べき等性: node_id / id を conflict_column として upsert。再実行可能。
"""

from __future__ import annotations

import argparse
import json
import logging
import os
import re
import sys
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import openpyxl
from dotenv import load_dotenv

# ---------------------------------------------------------------------------
# ロギング設定
# ---------------------------------------------------------------------------
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%Y-%m-%dT%H:%M:%S",
)
log = logging.getLogger("pace_import")

# ---------------------------------------------------------------------------
# .env 読み込み（プロジェクトルート → scripts/ の順）
# ---------------------------------------------------------------------------
_script_dir = Path(__file__).parent
load_dotenv(_script_dir.parent / ".env")
load_dotenv(_script_dir / ".env")

# ---------------------------------------------------------------------------
# Excel ファイルパス設定
# scripts/ の 2階層上 = Desktop/名称未設定フォルダ/
# ---------------------------------------------------------------------------
EXCEL_DIR = _script_dir.parent.parent

ASSESSMENT_FILES: dict[str, str] = {
    "F1_Acute":      "PACE_F1_Acute_FULL.xlsx",
    "F2_Chronic":    "PACE_F2_Chronic_FULL.xlsx",
    "F3_Performance":"PACE_F3_Performance_FULL.xlsx",
}

RTP_FILE      = "RTP_injury_specific_nodes_v2.2のコピー.xlsx"
MC_FILE       = "MC_tracking_nodes_v1.0のコピー.xlsx"
EXERCISE_FILE = "PACE_exercise_db_v3のコピー.xlsx"

# ---------------------------------------------------------------------------
# スキップするシート名パターン（INDEX / サマリー / 凡例 等）
# ---------------------------------------------------------------------------
_SKIP_SHEET_RE = re.compile(
    r"^(INDEX|凡例|サマリー|設計|フェーズ別)",
    re.IGNORECASE,
)


# ===========================================================================
# ユーティリティ
# ===========================================================================

def _normalize_key(raw: Any) -> str:
    """列名を正規化（小文字・記号をアンダースコアに統一）"""
    if raw is None:
        return ""
    s = str(raw).strip()
    # 改行を含むセルヘッダーを処理（例: "LR_Yes\n（臨床値）"）
    s = s.replace("\n", " ").replace("\r", " ")
    s = s.lower()
    # 括弧・スペース・記号・矢印 → アンダースコア
    s = re.sub(r"[（）()\s/・　\-→]+", "_", s)
    s = re.sub(r"_+", "_", s)
    return s.strip("_")


def _cell_str(v: Any) -> str:
    if v is None:
        return ""
    s = str(v).strip()
    return "" if s in ("nan", "None", "—", "-", "－") else s


def _parse_float(v: Any) -> float | None:
    if v is None:
        return None
    s = str(v).strip().replace(",", ".")
    # "Sn91%(SCAT5複合)" のような文字列から数値だけ取り出す試み
    m = re.match(r"^([+-]?\d+(?:\.\d+)?)", s)
    if m:
        try:
            return float(m.group(1))
        except ValueError:
            pass
    return None


def _parse_int(v: Any) -> int | None:
    f = _parse_float(v)
    return int(f) if f is not None else None


def _parse_json_list(raw: Any) -> list[str]:
    """
    セミコロン区切り / JSON 配列 / カンマ区切り を list[str] に変換。
    例: '!#AllExercise; !#AxialLoad'  → ['!#AllExercise', '!#AxialLoad']
    """
    if raw is None:
        return []
    s = str(raw).strip()
    if not s or s in ("—", "-", "nan", "None"):
        return []
    # JSON 配列として試みる
    if s.startswith("["):
        try:
            parsed = json.loads(s)
            if isinstance(parsed, list):
                return [str(x).strip() for x in parsed if str(x).strip()]
        except (json.JSONDecodeError, ValueError):
            pass
    # セミコロン区切り（Contraindication_Tags で使われる）
    if ";" in s:
        return [x.strip() for x in s.split(";") if x.strip()]
    # カンマ区切り
    return [x.strip() for x in s.split(",") if x.strip()]


def _open_workbook(path: Path) -> openpyxl.Workbook | None:
    if not path.exists():
        log.warning("ファイルが見つかりません: %s", path)
        return None
    try:
        return openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    except Exception as exc:
        log.warning("ファイルを開けませんでした %s: %s", path.name, exc)
        return None


# ===========================================================================
# Assessment (F1/F2/F3) パーサー
# ===========================================================================

# 列名 正規化キー → canonical フィールド名 のマッピング
# 優先度: 先に登録されたものが勝つ
_ASSESSMENT_COL_MAP: dict[str, str] = {
    # node_id
    "node_id":                   "node_id",
    "nodeid":                    "node_id",
    "node_id_固定":              "node_id",
    # phase
    "phase":                     "phase",
    "フェーズ":                  "phase",
    # category
    "category":                  "category",
    "カテゴリ":                  "category",
    # question_text
    "question_text":             "question_text",
    "質問文":                    "question_text",
    # target_axis
    "target_axis":               "target_axis",
    "対象axis":                  "target_axis",
    "axis":                      "target_axis",
    # lr_yes (κ補正後を優先)
    "lr_yes_sr_κ補正":          "lr_yes",
    "lr_yes_sr_κ補正後":        "lr_yes",
    "lr_yes_sr":                 "lr_yes",
    "lr_yes_臨床値":             "lr_yes",
    "lr_yes":                    "lr_yes",
    "lr+":                       "lr_yes",
    # lr_no
    "lr_no_clinical":            "lr_no",
    "lr_no除外力":               "lr_no",
    "lr_no":                     "lr_no",
    "lr-":                       "lr_no",
    # kappa
    "κ係数":                    "kappa",
    "kappa":                     "kappa",
    "cohen_kappa":               "kappa",
    # routing_rules (Routing_v4.3 等)
    "routing_v4.3":              "routing_rules",
    "routing_v4":                "routing_rules",
    "routing":                   "routing_rules",
    "routing_rules":             "routing_rules",
    # prescription_tags
    "prescription_tags":         "prescription_tags",
    "prescription":              "prescription_tags",
    # contraindication_tags
    "contraindication_tags":     "contraindication_tags",
    "contraindications":         "contraindication_tags",
    # time_decay_lambda
    "time_decay_λ":             "time_decay_lambda",
    "time_decay_lambda":         "time_decay_lambda",
    "time_decay":                "time_decay_lambda",
    "lambda":                    "time_decay_lambda",
    # evidence_level
    "エビデンスlv":              "evidence_level",
    "evidence_level":            "evidence_level",
    "エビデンスlevel":           "evidence_level",
    # sort_order (インポート順管理用; DB に入れない)
    "sort_order":                "sort_order",
}


def _map_columns(header_row: list[Any], alias_map: dict[str, str]) -> dict[int, str]:
    """
    {列インデックス: canonical名} を返す。
    同一 canonical は最初の列のみ登録（κ補正 vs 臨床値の優先制御）。
    ただし lr_yes だけは κ補正後（より右の列）を優先したいので
    後から登録し直す特殊処理を行う。
    """
    result: dict[int, str] = {}
    canonical_seen: dict[str, int] = {}  # canonical → 最初の列インデックス

    # LR_Yes は κ補正後列を使うため二段階で処理
    lr_yes_raw_idx: int | None = None    # 臨床値列
    lr_yes_sr_idx:  int | None = None    # κ補正後列

    for idx, cell_val in enumerate(header_row):
        key = _normalize_key(cell_val)
        if not key:
            continue
        canonical = alias_map.get(key)
        if canonical is None:
            continue

        # lr_yes 特別処理
        if canonical == "lr_yes":
            raw_key_ok = key in ("lr_yes_臨床値", "lr_yes", "lr+")
            sr_key_ok  = "sr" in key or "κ補正" in key
            if sr_key_ok:
                lr_yes_sr_idx = idx
            elif raw_key_ok and lr_yes_raw_idx is None:
                lr_yes_raw_idx = idx
            continue

        # 通常: 先着優先
        if canonical not in canonical_seen:
            result[idx] = canonical
            canonical_seen[canonical] = idx

    # lr_yes の最終決定: κ補正後 > 臨床値
    chosen_lr_yes = lr_yes_sr_idx if lr_yes_sr_idx is not None else lr_yes_raw_idx
    if chosen_lr_yes is not None:
        result[chosen_lr_yes] = "lr_yes"

    return result


def _find_header_row(rows: list[tuple], min_cells: int = 3) -> int:
    """
    ヘッダー行を探す。
    - Row 0 の唯一の非空セルが 'Node_ID' ならそのまま Row 0
    - そうでなければ Row 1 を確認（Excel の 1行目がタイトル行の場合）
    """
    for i, row in enumerate(rows[:5]):
        non_empty = [c for c in row if c is not None and str(c).strip() and str(c).strip() != "nan"]
        if len(non_empty) >= min_cells:
            return i
    return 0


def parse_assessment_sheet(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    file_type: str,
) -> list[dict[str, Any]]:
    """
    アセスメントシートを解析し DB 挿入可能な dict のリストを返す。
    ヘッダーは Row 0 (RTP/MC式) または Row 1 (F1/F2/F3式) に対応。
    """
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    header_idx = _find_header_row(rows, min_cells=4)
    header_row  = list(rows[header_idx])
    col_map     = _map_columns(header_row, _ASSESSMENT_COL_MAP)

    if not col_map:
        log.debug("  シート '%s': 認識できるヘッダーなし → スキップ", ws.title)
        return []

    records: list[dict[str, Any]] = []
    phase_from_sheet = _phase_from_sheet_name(ws.title)

    for row_idx, row in enumerate(rows[header_idx + 1:], start=header_idx + 2):
        if all(c is None or str(c).strip() in ("", "nan") for c in row):
            continue

        rec: dict[str, Any] = {"file_type": file_type}
        for col_idx, canonical in col_map.items():
            if col_idx >= len(row):
                continue
            raw = row[col_idx]
            if canonical in ("routing_rules", "prescription_tags", "contraindication_tags"):
                rec[canonical] = _parse_json_list(raw)
            elif canonical in ("lr_yes", "lr_no", "kappa", "time_decay_lambda"):
                rec[canonical] = _parse_float(raw)
            else:
                rec[canonical] = _cell_str(raw)

        # 必須: node_id または question_text が存在しなければスキップ
        if not rec.get("node_id") and not rec.get("question_text"):
            continue

        # node_id フォールバック生成
        if not rec.get("node_id"):
            rec["node_id"] = f"{file_type}_{ws.title}_{len(records)+1:03d}"

        # phase フォールバック
        if not rec.get("phase"):
            rec["phase"] = phase_from_sheet

        # question_text フォールバック
        if not rec.get("question_text"):
            rec["question_text"] = rec["node_id"]

        # デフォルト値補完 (DB NOT NULL)
        rec.setdefault("lr_yes", None)
        rec.setdefault("lr_no",  None)
        rec.setdefault("kappa",  None)
        if rec["lr_yes"] is None:
            rec["lr_yes"] = 1.0
        if rec["lr_no"] is None:
            rec["lr_no"] = 1.0
        if rec["kappa"] is None:
            rec["kappa"] = 0.0
        rec.setdefault("routing_rules",         [])
        rec.setdefault("prescription_tags",     [])
        rec.setdefault("contraindication_tags", [])
        rec.setdefault("time_decay_lambda",     0.0)

        # JSON 型を DB が期待する形式に変換
        rec["routing_rules"]         = json.dumps(rec["routing_rules"],         ensure_ascii=False)
        rec["prescription_tags"]     = json.dumps(rec["prescription_tags"],     ensure_ascii=False)
        rec["contraindication_tags"] = json.dumps(rec["contraindication_tags"], ensure_ascii=False)

        # sort_order を int に
        if "sort_order" in rec:
            rec["sort_order"] = _parse_int(rec["sort_order"]) or 0

        records.append(rec)

    return records


def _phase_from_sheet_name(sheet_name: str) -> str:
    """'1_RF_RedFlag' → 'RedFlag', '10_A2_負荷パターン' → 'Phase2' 等"""
    name = sheet_name.strip()
    m = re.match(r"^\d+_(.+)$", name)
    if m:
        remainder = m.group(1)
        return remainder.split("_")[0]
    return name[:30]


# ===========================================================================
# Exercise パーサー
# ===========================================================================

_EXERCISE_COL_MAP: dict[str, str] = {
    # category (列0 が category の場合が多い)
    "カテゴリ":                   "category",
    "category":                   "category",
    # phase
    "フェーズ":                   "phase",
    "phase":                      "phase",
    "phase_no":                   "phase",
    # name_ja
    "エクサイズ名":               "name_ja",
    "エクササイズ名":             "name_ja",
    "種目名":                     "name_ja",
    "種目":                       "name_ja",
    "exercise":                   "name_ja",
    "name_ja":                    "name_ja",
    "name":                       "name_ja",
    # target_axis
    "対象軸_部位":                "target_axis",
    "対象軸/部位":                "target_axis",
    "対象軸・部位":               "target_axis",
    "対象部位":                   "target_axis",
    "target_axis":                "target_axis",
    "axis":                       "target_axis",
    # sets
    "セット数":                   "sets",
    "sets":                       "sets",
    "set":                        "sets",
    # reps / time — 同一列に "8回" や "30秒" が混在することが多い
    "回数_時間":                  "reps_raw",   # 後処理で reps / time_sec に分割
    "回数・時間":                 "reps_raw",
    "reps":                       "reps_raw",
    # percent_1rm
    "rm目安":                     "percent_1rm",
    "rm_目安":                    "percent_1rm",
    "%1rm":                       "percent_1rm",
    "percent_1rm":                "percent_1rm",
    # rpe
    "rpe":                        "rpe",
    "強度":                       "rpe",
    # cues
    "主要キュー_実施ポイント":    "cues",
    "主要キュー・実施ポイント":   "cues",
    "主要キュー":                 "cues",
    "cues":                       "cues",
    # progressions
    "進行_バリエーション":        "progressions",
    "進行・バリエーション":       "progressions",
    "バリエーション":             "progressions",
    "progressions":               "progressions",
    # contraindication_tags
    "禁忌タグ":                   "contraindication_tags",
    "禁忌":                       "contraindication_tags",
    "contraindication_tags":      "contraindication_tags",
    "ci":                         "contraindication_tags",
}


def _split_reps_time(raw: Any) -> tuple[int | None, int | None]:
    """
    '8回' → (8, None)
    '30秒' → (None, 30)
    '100拍' → (100, None)
    '3×10回' → (10, None)
    """
    if raw is None:
        return None, None
    s = str(raw).strip()
    # 秒数
    m = re.search(r"(\d+)\s*秒", s)
    if m:
        return None, int(m.group(1))
    # 回・拍・step
    m = re.search(r"(\d+)\s*[回拍step]", s)
    if m:
        return int(m.group(1)), None
    # 純粋数値
    m = re.match(r"^(\d+(?:\.\d+)?)$", s)
    if m:
        return _parse_int(m.group(1)), None
    return None, None


def _extract_category_from_sheet_title(sheet_name: str) -> str:
    """
    '【ピラティス】 100種目' → 'ピラティス'
    'Phase 1 — 抑制・呼吸・代謝改善 （20種目）' → 'Phase 1'
    """
    m = re.search(r"【(.+?)】", sheet_name)
    if m:
        return m.group(1)
    # Phase N パターン
    m = re.match(r"(Phase\s*\d+)", sheet_name)
    if m:
        return m.group(1)
    return sheet_name[:20]


def parse_exercise_sheet(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    sheet_category: str,
) -> list[dict[str, Any]]:
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    # ヘッダー行を探す（少なくとも 2 列は実データ）
    header_idx = _find_header_row(rows, min_cells=2)
    header_row  = list(rows[header_idx])
    col_map     = _map_columns(header_row, _EXERCISE_COL_MAP)

    if not col_map:
        return []

    records: list[dict[str, Any]] = []

    for row in rows[header_idx + 1:]:
        if all(c is None or str(c).strip() in ("", "nan") for c in row):
            continue

        rec: dict[str, Any] = {}
        for col_idx, canonical in col_map.items():
            if col_idx >= len(row):
                continue
            raw = row[col_idx]
            if canonical == "contraindication_tags":
                rec[canonical] = _parse_json_list(raw)
            elif canonical in ("sets",):
                rec[canonical] = _parse_int(raw)
            elif canonical in ("rpe", "percent_1rm"):
                rec[canonical] = _parse_float(raw)
            elif canonical == "reps_raw":
                rec[canonical] = raw  # 後処理
            else:
                rec[canonical] = _cell_str(raw)

        # name_ja がなければスキップ
        if not rec.get("name_ja"):
            continue

        # reps / time_sec の分割処理
        reps_raw_val = rec.pop("reps_raw", None)
        reps, time_sec = _split_reps_time(reps_raw_val)
        rec["reps"]     = reps
        rec["time_sec"] = time_sec

        # category: 列データ優先、なければシート名から推定
        if not rec.get("category"):
            rec["category"] = sheet_category

        # phase: 列データ優先、なければ sheet_category から推定
        if not rec.get("phase"):
            m = re.search(r"\d+", sheet_category)
            rec["phase"] = m.group(0) if m else "1"

        # デフォルト補完
        rec.setdefault("name_en",               "")
        rec.setdefault("target_axis",           "")
        rec.setdefault("sets",                  1)
        rec.setdefault("cues",                  "")
        rec.setdefault("progressions",          "")
        rec.setdefault("contraindication_tags", [])

        # JSON 型変換
        rec["contraindication_tags"] = json.dumps(
            rec["contraindication_tags"], ensure_ascii=False
        )

        records.append(rec)

    return records


# ===========================================================================
# RTP Injury Nodes パーサー
# ===========================================================================

_RTP_COL_MAP: dict[str, str] = {
    "node_id":          "node_id",
    "id":               "node_id",
    # phase
    "フェーズ":         "phase",
    "phase":            "phase",
    # injury_type (シートタイトルから取る場合も多い)
    "injury_type":      "injury_type",
    "injury":           "injury_type",
    "カテゴリ":         "injury_type_col",  # カテゴリ列は injury_type の補助
    # gate_criteria (質問文 / criteria)
    "質問文":           "question_text",    # RTP ノードでは質問文が gate 内容
    "gate_criteria":    "gate_criteria",
    "criteria":         "gate_criteria",
    # lsi_target
    "lsi_target":       "lsi_target",
    "lsi":              "lsi_target",
    # test_battery
    "test_battery":     "test_battery",
    "tests":            "test_battery",
    # α_weight (参考情報として記録しない; gate_criteria に含める)
    "α_weight":         "alpha_weight",
    "α_weight_元値":   "alpha_weight",
}


def parse_rtp_sheet(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    injury_type_from_sheet: str,
) -> list[dict[str, Any]]:
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    header_idx = _find_header_row(rows, min_cells=2)
    header_row  = list(rows[header_idx])
    col_map     = _map_columns(header_row, _RTP_COL_MAP)

    if not col_map:
        return []

    records: list[dict[str, Any]] = []

    for row in rows[header_idx + 1:]:
        if all(c is None or str(c).strip() in ("", "nan") for c in row):
            continue

        rec: dict[str, Any] = {}
        for col_idx, canonical in col_map.items():
            if col_idx >= len(row):
                continue
            raw = row[col_idx]
            if canonical == "gate_criteria":
                # gate_criteria: JSON or 文字列
                s = _cell_str(raw)
                if s:
                    try:
                        rec[canonical] = json.loads(s)
                    except (json.JSONDecodeError, ValueError):
                        rec[canonical] = {"raw": s}
                else:
                    rec[canonical] = {}
            elif canonical == "test_battery":
                rec[canonical] = _parse_json_list(raw)
            elif canonical == "lsi_target":
                rec[canonical] = _parse_float(raw)
            elif canonical == "phase":
                # phase: テキスト型にする（"診断・分類" のような文字列もある）
                rec[canonical] = _cell_str(raw) or "1"
            elif canonical == "alpha_weight":
                pass  # gate_criteria に含めるか無視
            else:
                rec[canonical] = _cell_str(raw)

        if not rec.get("node_id"):
            rec["node_id"] = f"{injury_type_from_sheet}_{len(records)+1:03d}"

        # injury_type 決定
        if not rec.get("injury_type"):
            rec["injury_type"] = (
                rec.pop("injury_type_col", None) or injury_type_from_sheet
            )
        else:
            rec.pop("injury_type_col", None)

        # gate_criteria: question_text を組み込む
        if "question_text" in rec:
            qt = rec.pop("question_text")
            gate = rec.setdefault("gate_criteria", {})
            if isinstance(gate, dict) and "question" not in gate and qt:
                gate["question"] = qt

        # phase を int にキャスト（DB は INT型）
        phase_val = rec.get("phase", "1")
        phase_int = _parse_int(phase_val)
        if phase_int is None or not (1 <= phase_int <= 4):
            phase_int = 1  # 認識できないフェーズ名（"診断・分類"等）は 1 に
        rec["phase"] = phase_int

        rec.setdefault("gate_criteria", {})
        rec.setdefault("test_battery",  [])

        rec["gate_criteria"] = json.dumps(rec["gate_criteria"], ensure_ascii=False)
        rec["test_battery"]  = json.dumps(rec["test_battery"],  ensure_ascii=False)

        records.append(rec)

    return records


# ===========================================================================
# MC Tracking パーサー
# ===========================================================================

_MC_COL_MAP: dict[str, str] = {
    "node_id":              "node_id",
    "フェーズ":             "phase",
    "phase":                "phase",
    "カテゴリ":             "category",
    "質問文":               "question_text",
    "lr_yes_臨床値":        "lr_yes_raw",
    "lr_yes":               "lr_yes_raw",
    "lr_yes_sr_κ補正後":   "lr_yes",
    "lr_yes_sr":            "lr_yes",
    "κ係数":               "kappa",
    "lr_no_clinical":       "lr_no",
    "lr_no_κ非適用patch-c":"lr_no",
    "lr_no除外力":          "lr_no_excl",
    "対象axis":             "target_axis",
    "risk_flags":           "risk_flags",
}


def parse_mc_sheet(ws: openpyxl.worksheet.worksheet.Worksheet) -> list[dict[str, Any]]:
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    header_idx = _find_header_row(rows, min_cells=3)
    header_row  = list(rows[header_idx])
    col_map     = _map_columns(header_row, _MC_COL_MAP)

    if not col_map:
        return []

    records: list[dict[str, Any]] = []

    for row in rows[header_idx + 1:]:
        if all(c is None or str(c).strip() in ("", "nan") for c in row):
            continue

        rec: dict[str, Any] = {}
        for col_idx, canonical in col_map.items():
            if col_idx >= len(row):
                continue
            raw = row[col_idx]
            if canonical in ("lr_yes", "lr_yes_raw", "lr_no", "lr_no_excl", "kappa"):
                rec[canonical] = _parse_float(raw)
            elif canonical == "risk_flags":
                rec[canonical] = _parse_json_list(raw)
            else:
                rec[canonical] = _cell_str(raw)

        if not rec.get("node_id"):
            continue

        # lr_yes: κ補正後優先
        if rec.get("lr_yes") is None and rec.get("lr_yes_raw") is not None:
            rec["lr_yes"] = rec["lr_yes_raw"]
        rec.pop("lr_yes_raw", None)
        rec.pop("lr_no_excl", None)

        # デフォルト値
        if rec.get("lr_yes") is None:
            rec["lr_yes"] = 1.0
        if rec.get("lr_no") is None:
            rec["lr_no"] = 1.0
        if rec.get("kappa") is None:
            rec["kappa"] = 0.0

        rec.setdefault("risk_flags", [])
        rec["risk_flags"] = json.dumps(rec["risk_flags"], ensure_ascii=False)

        records.append(rec)

    return records


# ===========================================================================
# Alpha Chains パーサー（MC ワークブックの αチェーン定義_MC シート）
# ===========================================================================

_ALPHA_CHAIN_COL_MAP: dict[str, str] = {
    "chain_id":               "chain_id",
    "chain_id_patcha":        "chain_id",
    # "Chain名" は正規化すると "chain名" になる
    "chain名":                "chain_name",
    "chain_name":             "chain_name",
    # "構成ノード（→α値）" は正規化すると "構成ノード_α値" になる
    "構成ノード_α値":        "nodes_raw",
    "構成ノード_α値_":       "nodes_raw",
    # "連鎖根拠" は正規化後もそのまま
    "連鎖根拠":               "causal_reasoning",
    "causal_reasoning":       "causal_reasoning",
}


def parse_alpha_chain_sheet(ws: openpyxl.worksheet.worksheet.Worksheet) -> list[dict[str, Any]]:
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    header_idx = _find_header_row(rows, min_cells=2)
    header_row  = list(rows[header_idx])
    col_map     = _map_columns(header_row, _ALPHA_CHAIN_COL_MAP)

    if not col_map:
        return []

    records: list[dict[str, Any]] = []

    for row in rows[header_idx + 1:]:
        if all(c is None or str(c).strip() in ("", "nan") for c in row):
            continue

        rec: dict[str, Any] = {}
        for col_idx, canonical in col_map.items():
            if col_idx >= len(row):
                continue
            rec[canonical] = _cell_str(row[col_idx])

        if not rec.get("chain_id"):
            continue

        # nodes_raw → JSON
        # 例: "MC_PH_001(1.0)→MC_IR_001(1.0)→MC_TR_004(1.0)"
        nodes_raw = rec.pop("nodes_raw", "")
        nodes: list[dict[str, Any]] = []
        for part in re.split(r"[→\->]+", nodes_raw):
            part = part.strip()
            nm = re.match(r"^(.+?)\(([0-9.]+)\)$", part)
            if nm:
                nodes.append({"node_id": nm.group(1), "alpha": float(nm.group(2))})
            elif part:
                nodes.append({"node_id": part, "alpha": 1.0})
        rec["nodes_json"] = json.dumps(nodes, ensure_ascii=False)

        rec.setdefault("causal_reasoning", "")

        records.append(rec)

    return records


# ===========================================================================
# Supabase クライアント & upsert
# ===========================================================================

def get_supabase_client():
    url = os.environ.get("SUPABASE_URL")
    key = os.environ.get("SUPABASE_SERVICE_ROLE_KEY")
    if not url or not key:
        log.error(
            "SUPABASE_URL と SUPABASE_SERVICE_ROLE_KEY が設定されていません。"
            " .env を作成するか環境変数をエクスポートしてください。"
        )
        sys.exit(1)
    from supabase import create_client
    return create_client(url, key)


def upsert_records(
    client,
    table: str,
    records: list[dict[str, Any]],
    conflict_column: str,
    dry_run: bool,
) -> int:
    """records を table に upsert する。成功件数を返す。"""
    if not records:
        return 0
    if dry_run:
        log.info("  [DRY-RUN] %s へ %d 件を upsert 予定", table, len(records))
        for r in records[:2]:
            log.info("    例: %s", {k: v for k, v in list(r.items())[:6]})
        return len(records)

    CHUNK = 200  # Supabase の HTTP ペイロード制限を避けるためチャンク分割
    total = 0
    for i in range(0, len(records), CHUNK):
        chunk = records[i : i + CHUNK]
        try:
            client.table(table).upsert(
                chunk, on_conflict=conflict_column
            ).execute()
            total += len(chunk)
        except Exception as exc:
            log.error("  [ERROR] %s upsert 失敗 (rows %d-%d): %s", table, i, i + len(chunk), exc)
    return total


def log_import_result(
    client,
    table: str,
    file_name: str,
    sheet_name: str,
    rows_processed: int,
    dry_run: bool,
) -> None:
    """import_logs テーブルにインポート結果を記録する（テーブルが存在する場合のみ）。"""
    if dry_run or client is None:
        return
    try:
        client.table("import_logs").insert({
            "file_name":      file_name,
            "sheet_name":     sheet_name,
            "target_table":   table,
            "rows_processed": rows_processed,
            "imported_at":    datetime.now(timezone.utc).isoformat(),
        }).execute()
    except Exception:
        pass  # import_logs テーブルが未作成の場合はサイレントスキップ


# ===========================================================================
# バリデーション
# ===========================================================================

_ROUTING_SKIP_TOKENS = frozenset({
    # Excel で "Always" はルーティングを省略なく全ノードに適用することを意味する説明文
    "always", "if", "p0_001", "then", "or", "and", "when",
})


def validate_assessment_records(
    records: list[dict[str, Any]],
    all_node_ids: set[str],
) -> list[str]:
    """バリデーションエラー文字列のリストを返す。"""
    errors: list[str] = []
    for rec in records:
        nid = rec.get("node_id", "?")
        lr_yes = rec.get("lr_yes")
        lr_no  = rec.get("lr_no")
        # LR 値チェック
        if lr_yes is not None and lr_yes <= 0:
            errors.append(f"{nid}: lr_yes={lr_yes} が 0 以下 (無効な LR 値)")
        if lr_no is not None and lr_no <= 0:
            errors.append(f"{nid}: lr_no={lr_no} が 0 以下 (無効な LR 値)")
        # Routing ルーティング先存在チェック（同バッチ内で確認）
        routing = json.loads(rec.get("routing_rules", "[]"))
        for target in routing:
            # "Always（全ファイル）" のような説明トークンと
            # 条件式リテラル ("If P0_001=局所力学" 等) はスキップ
            t_lower = target.lower().split("（")[0].strip()
            if t_lower in _ROUTING_SKIP_TOKENS:
                continue
            # 条件式（スペースや日本語を含む）はスキップ
            if " " in target or "=" in target or "（" in target:
                continue
            if re.match(r"^[A-Za-z0-9_\-\.]+$", target) and target not in all_node_ids:
                errors.append(f"{nid}: routing 先 '{target}' が今回のバッチに存在しない")
    return errors


# ===========================================================================
# インポートルーティン
# ===========================================================================

def import_assessment_files(client, dry_run: bool) -> None:
    all_node_ids: set[str] = set()

    for label, filename in ASSESSMENT_FILES.items():
        path = EXCEL_DIR / filename
        log.info("\n[Assessment] %s  ←  %s", label, filename)
        wb = _open_workbook(path)
        if wb is None:
            continue

        file_records: list[dict[str, Any]] = []

        for sheet_name in wb.sheetnames:
            if _SKIP_SHEET_RE.match(sheet_name):
                log.debug("  シート '%s': スキップ（INDEX/サマリー系）", sheet_name)
                continue
            try:
                ws = wb[sheet_name]
                records = parse_assessment_sheet(ws, label)
                log.info("  シート '%-30s': %3d ノード", sheet_name, len(records))
                file_records.extend(records)
            except Exception as exc:
                log.error("  [ERROR] シート '%s': %s", sheet_name, exc)

        wb.close()

        # バリデーション
        batch_ids = {r["node_id"] for r in file_records}
        all_node_ids |= batch_ids
        errors = validate_assessment_records(file_records, all_node_ids)
        if errors:
            log.warning("  バリデーション警告 (%d 件):", len(errors))
            for e in errors[:10]:
                log.warning("    %s", e)

        count = upsert_records(client, "assessment_nodes", file_records, "node_id", dry_run)
        log.info("  → upsert 完了: %d 件 (assessment_nodes)", count)
        log_import_result(client, "assessment_nodes", filename, "*", count, dry_run)


def import_exercise_file(client, dry_run: bool) -> None:
    path = EXCEL_DIR / EXERCISE_FILE
    log.info("\n[Exercise DB]  ←  %s", EXERCISE_FILE)
    wb = _open_workbook(path)
    if wb is None:
        return

    all_records: list[dict[str, Any]] = []

    for sheet_name in wb.sheetnames:
        if _SKIP_SHEET_RE.match(sheet_name):
            continue
        # 全エクサイズ一覧シートはスキップ（個別シートで取り込む）
        if "全エクサイズ" in sheet_name or "一覧" in sheet_name:
            continue
        try:
            ws = wb[sheet_name]
            category = _extract_category_from_sheet_title(sheet_name)
            records  = parse_exercise_sheet(ws, category)
            log.info("  シート '%-30s' (cat=%s): %3d 種目", sheet_name, category, len(records))
            all_records.extend(records)
        except Exception as exc:
            log.error("  [ERROR] シート '%s': %s", sheet_name, exc)

    wb.close()

    # exercises は (category, phase, name_ja) の複合ユニーク制約で upsert する
    # DB 側に UNIQUE(category, phase, name_ja) が必要（migration SQL 参照）
    count = upsert_records(
        client, "exercises", all_records,
        "category,phase,name_ja", dry_run
    )
    log.info("  → upsert 完了: %d 件 (exercises)", count)
    log_import_result(client, "exercises", EXERCISE_FILE, "*", count, dry_run)


def import_rtp_file(client, dry_run: bool) -> None:
    path = EXCEL_DIR / RTP_FILE
    log.info("\n[RTP Nodes]  ←  %s", RTP_FILE)
    wb = _open_workbook(path)
    if wb is None:
        return

    all_records: list[dict[str, Any]] = []

    for sheet_name in wb.sheetnames:
        if _SKIP_SHEET_RE.match(sheet_name):
            continue
        try:
            ws = wb[sheet_name]
            # injury_type: シート名から推定（例: "ACL（再建術・保存療法）" → "ACL"）
            injury_m = re.match(r"^([A-Za-z\u3000-\u9fff\u4e00-\u9fff]+)", sheet_name)
            injury_type = injury_m.group(1).strip() if injury_m else sheet_name[:20]
            records = parse_rtp_sheet(ws, injury_type)
            log.info("  シート '%-30s': %3d ノード", sheet_name, len(records))
            all_records.extend(records)
        except Exception as exc:
            log.error("  [ERROR] シート '%s': %s", sheet_name, exc)

    wb.close()

    count = upsert_records(client, "rtp_injury_nodes", all_records, "node_id", dry_run)
    log.info("  → upsert 完了: %d 件 (rtp_injury_nodes)", count)
    log_import_result(client, "rtp_injury_nodes", RTP_FILE, "*", count, dry_run)


def import_mc_file(client, dry_run: bool) -> None:
    path = EXCEL_DIR / MC_FILE
    log.info("\n[MC Tracking]  ←  %s", MC_FILE)
    wb = _open_workbook(path)
    if wb is None:
        return

    for sheet_name in wb.sheetnames:
        if _SKIP_SHEET_RE.match(sheet_name):
            continue

        # αチェーン定義シート → alpha_chains テーブルへ
        if "αチェーン" in sheet_name or "alpha_chain" in sheet_name.lower():
            try:
                ws = wb[sheet_name]
                records = parse_alpha_chain_sheet(ws)
                log.info("  シート '%-30s': %3d αチェーン", sheet_name, len(records))
                count = upsert_records(client, "alpha_chains", records, "chain_id", dry_run)
                log.info("  → upsert 完了: %d 件 (alpha_chains)", count)
                log_import_result(client, "alpha_chains", MC_FILE, sheet_name, count, dry_run)
            except Exception as exc:
                log.error("  [ERROR] シート '%s': %s", sheet_name, exc)
            continue

        # MC ノードシート → mc_tracking_nodes テーブルへ
        try:
            ws = wb[sheet_name]
            records = parse_mc_sheet(ws)
            log.info("  シート '%-30s': %3d MC ノード", sheet_name, len(records))
            count = upsert_records(client, "mc_tracking_nodes", records, "node_id", dry_run)
            log.info("  → upsert 完了: %d 件 (mc_tracking_nodes)", count)
            log_import_result(client, "mc_tracking_nodes", MC_FILE, sheet_name, count, dry_run)
        except Exception as exc:
            log.error("  [ERROR] シート '%s': %s", sheet_name, exc)

    wb.close()


# ===========================================================================
# エントリポイント
# ===========================================================================

def main() -> None:
    parser = argparse.ArgumentParser(
        description="PACE Platform — Excel → Supabase インポートスクリプト"
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Supabase への書き込みをせずプレビューのみ表示",
    )
    parser.add_argument(
        "--only",
        choices=["assessment", "exercises", "rtp", "mc"],
        help="特定カテゴリのみインポート（デフォルト: 全て）",
    )
    parser.add_argument(
        "--verbose", "-v",
        action="store_true",
        help="DEBUG ログを表示",
    )
    args = parser.parse_args()

    if args.verbose:
        logging.getLogger().setLevel(logging.DEBUG)

    if args.dry_run:
        log.info("=" * 60)
        log.info("DRY-RUN モード — Supabase へのデータ書き込みは行いません")
        log.info("=" * 60)
        client = None
    else:
        client = get_supabase_client()

    only = args.only

    if only is None or only == "assessment":
        import_assessment_files(client, args.dry_run)

    if only is None or only == "exercises":
        import_exercise_file(client, args.dry_run)

    if only is None or only == "rtp":
        import_rtp_file(client, args.dry_run)

    if only is None or only == "mc":
        import_mc_file(client, args.dry_run)

    log.info("\n完了。")


if __name__ == "__main__":
    main()
